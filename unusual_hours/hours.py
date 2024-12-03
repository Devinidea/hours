import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter

def generate_weekdays(year, month):
    all_dates = []
    
    start_date = f'{year}-{month:02d}-01'
    # 遍历该月的每一天
    end_date = f'{year}-{month:02d}-{pd.Period(start_date).days_in_month}'  # 获取该月最后一天
    
    # 遍历该月的每一天
    for single_date in pd.date_range(start=start_date, end=end_date):
        # 判断是否是工作日（周一到周五）
        if single_date.weekday() < 5:  # 0=周一, 1=周二, ..., 4=周五
            all_dates.append(single_date)

    return all_dates

def save_to_excel(dates, names, filename):
    # 将日期格式转换为 'YYYY-MM-DD'
    dates_formatted = [date.strftime('%Y-%m-%d') for date in dates]
    df = pd.DataFrame({'日期 Date': dates_formatted, '姓名 Name': names, '组别 Team': ['']*len(dates), '岗外作业类型 Type of loss time': ['']*len(dates)})
    df.to_excel(filename, index=False)
    
    # 调整列宽并设置日期格式
    adjust_column_width_and_format(filename)

def adjust_column_width_and_format(filename):
    # 使用 openpyxl 加载工作簿
    wb = load_workbook(filename)
    ws = wb.active
    
    # 创建日期格式样式
    date_style = NamedStyle(name="datetime", number_format="YYYY-MM-DD")
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
                # 设置日期列的格式
                if column == 'A':  # 假设日期列是第一列
                    cell.style = date_style
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(filename)

if __name__ == "__main__":
    year = int(input("请输入年份（例如 2024）："))
    month = int(input("请输入月份（1-12，例如 10）："))
    
    weekdays = generate_weekdays(year, month)
    
    # 生成与日期长度一致的姓名列表
    name = input("请输入姓名：")
    names = [name] * len(weekdays)
    
    # 保存到Excel文件
    save_to_excel(weekdays, names, '异常工时.xlsx')